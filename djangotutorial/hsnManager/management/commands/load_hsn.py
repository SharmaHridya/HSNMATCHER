"""
Load the HSN master list from a single Excel sheet.

Usage:
    python manage.py load_hsn <master_xlsx>

The workbook should contain one sheet with:

Column A -> HSN Code
Column B -> Description
Column C -> GST Rate

The table is wiped and reloaded every time.
"""

import re
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from hsnManager.models import HSNCode


def normalize_code(raw):
    if raw is None:
        return None

    digits = re.sub(r"[^0-9]", "", str(raw))

    if not digits:
        return None

    return digits.zfill(8)


class Command(BaseCommand):
    help = "Load HSN master data into the database."

    def add_arguments(self, parser):
        parser.add_argument(
            "master",
            help="Path to the master Excel file",
        )

    def handle(self, *args, **options):
        try:
            workbook = openpyxl.load_workbook(
                options["master"],
                read_only=True,
            )
        except Exception as e:
            raise CommandError(f"Could not open workbook: {e}")

        worksheet = workbook[workbook.sheetnames[0]]

        rows = []
        seen = set()
        skipped = 0

        for raw_code, raw_desc, raw_rate in worksheet.iter_rows(
            min_row=2,
            max_col=3,
            values_only=True,
        ):
            code = normalize_code(raw_code)
            desc = " ".join(str(raw_desc).split()) if raw_desc else ""

            if not code or not desc or len(code) > 8:
                skipped += 1
                continue

            if code in seen:
                skipped += 1
                continue

            seen.add(code)

            rows.append(
                HSNCode(
                    code=code,
                    description=desc,
                    gst_rate=raw_rate,
                    code_type="HSN",
                )
            )

        with transaction.atomic():
            deleted, _ = HSNCode.objects.all().delete()
            HSNCode.objects.bulk_create(rows, batch_size=1000)

        self.stdout.write(
            self.style.SUCCESS(
                f"Loaded {len(rows)} codes; "
                f"skipped {skipped} bad/duplicate rows; "
                f"replaced {deleted} existing records."
            )
        )