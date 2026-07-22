"""
Load the HSN/SAC master list (and optionally GST rates) into the database.

Usage:
    python manage.py load_hsn <master_xlsx> [--rates <rates_xlsx>]

The master workbook is expected to have sheets HSN_MSTR (HSN_CD, HSN_Description)
and SAC_MSTR (SAC_CD, SAC_Description) — the format of the GST portal's
HSN/SAC master download.

The optional rates workbook is expected to have a 'GST' sheet with columns
(Sl. No., HSN Code No., Name of Commodity, Chapter No., Sch., GST %) starting
around row 10 — dotted codes ('0101.00.00'), fractional rates (0.18), 'NIL'
meaning 0%, and '--' meaning no rate listed.

The load is idempotent: it wipes and reloads the table inside a transaction.
"""

import re
from decimal import Decimal

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from hsnManager.models import HSNCode


def normalize_code(raw):
    """'0508 .00.10\\xa0' -> '05080010'; returns None if not a real code."""
    if raw is None:
        return None
    digits = re.sub(r"[^0-9]", "", str(raw))
    return digits or None


def parse_rate(raw):
    """0.18 -> Decimal('18'); 'NIL' -> Decimal('0'); '--'/None -> None."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return Decimal(str(round(raw * 100, 2)))
    text = str(raw).strip().upper()
    if text == "NIL":
        return Decimal("0")
    return None  # '--', '---', blanks: no rate listed


class Command(BaseCommand):
    help = "Load the HSN/SAC master list (and optionally GST rates) into the DB."

    def add_arguments(self, parser):
        parser.add_argument("master", help="Path to the HSN/SAC master .xlsx")
        parser.add_argument(
            "--rates",
            help="Path to a GST rates .xlsx to merge rates from (goods only)",
        )

    def handle(self, *args, **options):
        try:
            master = openpyxl.load_workbook(options["master"], read_only=True)
        except Exception as e:
            raise CommandError(f"Could not open master workbook: {e}")

        rows = []  # (code, description, code_type)
        seen = set()
        skipped = 0
        for sheet, code_type in (("HSN_MSTR", "HSN"), ("SAC_MSTR", "SAC")):
            if sheet not in master.sheetnames:
                raise CommandError(f"Master workbook has no '{sheet}' sheet")
            for raw_code, raw_desc in master[sheet].iter_rows(
                min_row=2, max_col=2, values_only=True
            ):
                code = normalize_code(raw_code)
                desc = " ".join(str(raw_desc).split()) if raw_desc else ""
                if not code or not desc or len(code) > 8:
                    skipped += 1
                    continue
                if code in seen:  # first occurrence wins
                    skipped += 1
                    continue
                seen.add(code)
                rows.append((code, desc, code_type))

        rates = {}
        if options["rates"]:
            rates = self.load_rates(options["rates"])

        matched = 0
        objs = []
        for code, desc, code_type in rows:
            # Master codes are 2-8 digits; the rate schedule keys on 8.
            # '0101' padded to '01010000' matches rate row '0101.00.00'.
            rate = rates.get(code) or rates.get(code.ljust(8, "0"))
            if rate is not None:
                matched += 1
            objs.append(
                HSNCode(code=code, description=desc, gst_rate=rate, code_type=code_type)
            )

        with transaction.atomic():
            deleted, _ = HSNCode.objects.all().delete()
            HSNCode.objects.bulk_create(objs, batch_size=1000)

        self.stdout.write(
            self.style.SUCCESS(
                f"Loaded {len(objs)} codes "
                f"({sum(1 for r in rows if r[2] == 'HSN')} HSN, "
                f"{sum(1 for r in rows if r[2] == 'SAC')} SAC); "
                f"skipped {skipped} bad/duplicate rows; "
                f"replaced {deleted} existing; "
                f"rates matched for {matched} codes."
            )
        )

    def load_rates(self, path):
        """Return {8-digit code: Decimal rate} from the rates workbook."""
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
        except Exception as e:
            raise CommandError(f"Could not open rates workbook: {e}")
        if "GST" not in wb.sheetnames:
            raise CommandError("Rates workbook has no 'GST' sheet")

        rates = {}
        conflicts = 0
        for row in wb["GST"].iter_rows(min_row=10, values_only=True):
            code = normalize_code(row[2])
            rate = parse_rate(row[6])
            if not code or rate is None:
                continue
            if code in rates and rates[code] != rate:
                conflicts += 1  # keep the first listing
                continue
            rates[code] = rate

        self.stdout.write(
            f"Rates file: {len(rates)} coded rates parsed"
            + (f", {conflicts} conflicting duplicates ignored" if conflicts else "")
        )
        return rates
